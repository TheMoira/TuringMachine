import turing_utils.turing_draft as turing
from openpyxl import Workbook, styles as ops, load_workbook
table_corner_row = 4
table_corner_column = 1

def generate_xlsx_file(name, alphabet, number_of_states, empty_mark='#'):
    filename = f"{name}_machine_instructions"

    if empty_mark not in alphabet:
        alphabet.insert(0,empty_mark)

    workbook = Workbook()
    sheet = workbook.active

    bold_font = ops.Font(bold=True)
    square_border = ops.Border(ops.Side(style="hair"),ops.Side(style="hair"),
                               ops.Side(style="hair"),ops.Side(style="hair"))
    center_alignment = ops.Alignment(horizontal="center", vertical="center")
    colors = {"value": "fffbf2", "state": "fcf0ff",
              "step": "e3eeff"}

    sheet.cell(table_corner_row, table_corner_column, 'ALPHABET')
    for state_number in range(number_of_states):
        sheet.cell(table_corner_row, table_corner_column + state_number + 1, f'q{state_number}')

    for idx in range(len(alphabet)):
        current = 3 * idx + table_corner_row + 1
        sheet.cell(row=current, column=table_corner_column).font = bold_font
        sheet.cell(row=current, column=table_corner_column).value = alphabet[idx]
        sheet.merge_cells(start_row=current, start_column=table_corner_column,
                          end_row=current+2, end_column=table_corner_column)
        sheet.cell(row=current, column=table_corner_column).alignment = center_alignment

    sheet.protection.sheet = True

    for row_number in range(table_corner_row + 1, len(alphabet)*3 + table_corner_row + 1):
        if row_number%3 == 1:
            color = colors["step"]
        elif row_number%3 == 2:
            color = colors["value"]
        else:
            color = colors["state"]
        for col in range(table_corner_column + 1, number_of_states + table_corner_column + 1):
            sheet.cell(row=row_number, column=col).protection = ops.Protection(locked=False)
            sheet.cell(row=row_number, column=col).fill = ops.PatternFill("solid", fgColor=color)
            sheet.cell(row=row_number, column=col).border = square_border

    sheet["A1"] = f"Instrukcja {name}"
    sheet["A1"].font = bold_font
    sheet.merge_cells("A1:B1")
    sheet["A2"] = "value"
    sheet["A2"].fill = ops.PatternFill("solid", fgColor=colors["value"])
    sheet["B2"] = "state"
    sheet["B2"].fill = ops.PatternFill("solid", fgColor=colors["state"])
    sheet["C2"] = "step"
    sheet["C2"].fill = ops.PatternFill("solid", fgColor=colors["step"])

    sheet.freeze_panes = f'A{table_corner_row+1}'

    workbook.save(f"{filename}.xlsx")

def generate_instruction_from_xlsx_file(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    states_row = sheet[table_corner_row]
    number_of_states = len(states_row) - 1
    alphabet_len = (sheet.max_row - table_corner_row) // 3
    instructions = []

    for state_number in range(number_of_states):
        state_position_col = state_number + table_corner_column + 1
        for value_number in range(alphabet_len):
            value_position_row = value_number * 3 + table_corner_row + 1
            in_val = sheet.cell(value_position_row, table_corner_column).value
            in_state = sheet.cell(table_corner_row, state_position_col).value
            out_val = sheet.cell(value_position_row, state_position_col).value
            out_state = sheet.cell(value_position_row + 1, state_position_col).value
            step = sheet.cell(value_position_row + 2, state_position_col).value
            instr = (in_val, in_state, out_val, out_state, step)
            instructions.append(turing.Instruction(instr))
    return instructions



if __name__ == '__main__':
    alp = ['#', '0', '1']
    alp2 = ['a', 'b', 'c', 'd']
    name = 'test2'
    name2 = 'test1'
    nr = 2
    nr2 = 5

    # generate_xlsx_file(name,alp,nr)
    # generate_xlsx_file(name2,alp2,nr2)

    # instructions = generate_instruction_from_xlsx_file('test2_machine_instructions.xlsx')
    # for instr in instructions:
    #     print(instr)




