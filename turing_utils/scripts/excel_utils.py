import turing_utils.scripts.turing_draft as turing
from openpyxl import Workbook, styles as ops, load_workbook
import sys

table_corner_row = 4
table_corner_column = 1


def generate_xlsx_file(name, alphabet, number_of_states, path = './', empty_mark='#', return_only_workbook = False):
    filename = f"{name}_machine_instructions"

    if empty_mark not in alphabet:
        alphabet.insert(0,empty_mark)

    workbook = Workbook()
    sheet = workbook.active

    bold_font = ops.Font(bold=True)
    square_border = ops.Border(ops.Side(style="hair"),ops.Side(style="hair"),
                               ops.Side(style="hair"),ops.Side(style="hair"))
    center_alignment = ops.Alignment(horizontal="center", vertical="center")
    colors = {"value": "fffbf2", "state": "fcf0ff",
              "step": "e3eeff"}

    sheet.cell(table_corner_row, table_corner_column, 'ALPHABET')
    for state_number in range(number_of_states):
        sheet.cell(table_corner_row, table_corner_column + state_number + 1, f'q{state_number}')

    for idx in range(len(alphabet)):
        current = 3 * idx + table_corner_row + 1
        sheet.cell(row=current, column=table_corner_column).font = bold_font
        sheet.cell(row=current, column=table_corner_column).value = alphabet[idx]
        sheet.merge_cells(start_row=current, start_column=table_corner_column,
                          end_row=current+2, end_column=table_corner_column)
        sheet.cell(row=current, column=table_corner_column).alignment = center_alignment

    sheet.protection.sheet = True

    for row_number in range(table_corner_row + 1, len(alphabet)*3 + table_corner_row + 1):
        if row_number%3 == 1:
            color = colors["step"]
        elif row_number%3 == 2:
            color = colors["value"]
        else:
            color = colors["state"]
        for col in range(table_corner_column + 1, number_of_states + table_corner_column + 1):
            sheet.cell(row=row_number, column=col).protection = ops.Protection(locked=False)
            sheet.cell(row=row_number, column=col).fill = ops.PatternFill("solid", fgColor=color)
            sheet.cell(row=row_number, column=col).border = square_border

    sheet["A1"] = f"Instrukcja {name}"
    sheet["A1"].font = bold_font
    sheet.merge_cells("A1:B1")
    sheet["A2"] = "value"
    sheet["A2"].fill = ops.PatternFill("solid", fgColor=colors["value"])
    sheet["B2"] = "state"
    sheet["B2"].fill = ops.PatternFill("solid", fgColor=colors["state"])
    sheet["C2"] = "step"
    sheet["C2"].fill = ops.PatternFill("solid", fgColor=colors["step"])

    sheet.freeze_panes = f'A{table_corner_row+1}'

    if return_only_workbook:
        return workbook

    workbook.save(f"{path}{filename}.xlsx")
    return f"{path}{filename}.xlsx"


def generate_instructions_from_xlsx_file(filename, only_as_tuples = False, fileout = None, examples = None, ready_workbook = None):
    if not ready_workbook:
        workbook = load_workbook(filename)
    else:
        workbook = ready_workbook
    sheet = workbook.active
    states_row = sheet[table_corner_row]
    number_of_states = len(states_row) - 1
    alphabet_len = (sheet.max_row - table_corner_row) // 3
    instructions = []
    original_stdout = sys.stdout
    if(fileout):
        sys.stdout = open(fileout, 'w')
    print('Instructions:')
    for state_number in range(number_of_states):
        state_position_col = state_number + table_corner_column + 1
        for value_number in range(alphabet_len):
            value_position_row = value_number * 3 + table_corner_row + 1
            in_val = sheet.cell(value_position_row, table_corner_column).value
            in_state = sheet.cell(table_corner_row, state_position_col).value
            out_val = sheet.cell(value_position_row, state_position_col).value
            out_state = sheet.cell(value_position_row + 1, state_position_col).value
            step = sheet.cell(value_position_row + 2, state_position_col).value
            if only_as_tuples:
                instr = f'({in_val},{in_state},{out_val},{out_state},{step})'
                instructions.append(instr)
                print(instr)
            else:
                instr = (in_val, in_state, out_val, out_state, step)
                instructions.append(turing.Instruction(instr))
    if(examples):
        print('Examples:')
        for e in examples:
            for l in e:
                print(l, end=' ')
            print()
    sys.stdout = original_stdout
    return instructions



if __name__ == '__main__':
    alp = ['#', '0', '1']
    alp2 = ['a', 'b', 'c', 'd']
    alp3 = ['a', 'b']

    name = 'test2'
    name2 = 'test1'
    name3 = 'double_content'
    
    nr = 2
    nr2 = 5
    nr3 = 9

    examples3 = [['#', 'a', 'a', 'b', '#']]

    # generate_xlsx_file(name,alp,nr)
    # generate_xlsx_file(name2,alp2,nr2)
    # generate_xlsx_file(name3,alp3,nr3)

    instructions = generate_instructions_from_xlsx_file(f'{name3}_machine_instructions.xlsx', True, 'test_files/input_2.txt', examples3)
    # for instr in instructions:
    #     print(instr)




